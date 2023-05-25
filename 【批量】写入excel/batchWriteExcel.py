from openpyxl import Workbook
import os

rootPath = r"D:\myspace\anki-all\额外的东西"


for file in os.listdir(rootPath):
    a = []
    b = []
    suff_name = os.path.splitext(file)[1]  # 获取文件后缀
    if suff_name == '.txt':
        file_name = os.path.splitext(file)[0]  # 获取文件名称
        path = os.path.join(rootPath + '//' + file_name+'.txt')  # 获取文件路径
        path2 = os.path.join(rootPath + '//' + file_name+'.xlsx')  # 获取文件路径
        with open(path, "rb") as f:
            for index, line in enumerate(f.readlines()):  # 打开后逐行读取
                if index % 2 == 0:
                    a.append(line)
                else:
                    b.append(line)
            workbook = Workbook()
            sheet0 = workbook.create_sheet(index=0)  # 创建sheet0
            sheet0.column_dimensions['A'].width = 15  # 设置A列宽度
            sheet0.column_dimensions['B'].width = 22  # 设置B列宽度
            sheet0.column_dimensions['C'].width = 15  # 设置C列宽度
            sheet0.column_dimensions['D'].width = 15  # 设置D列宽度
            # 循环写入数据，居中对齐
            for i in range(len(a)):
                sheet0.cell(i+1, 1).value = a[i]  # 写入数据

            for j in range(len(b)):
                sheet0.cell(j+1, 2).value = b[j]  # 写入数据

            workbook.save(path2)  # 保存文件
