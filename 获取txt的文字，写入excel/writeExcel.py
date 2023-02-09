from openpyxl import Workbook
# 首先给出文件路径
resultpath = r"C:\Users\geniusShi\Desktop\tran分句版本.txt"
a = []
b = []


with open(resultpath, "rb") as f:
    for index, line in enumerate(f.readlines()):  # 打开后逐行读取
        if index % 2 == 0:
            a.append(line)
        else:
            b.append(line)
        



# 说明：需要写入的是二维列表target_data
# 将数据写入excel表格
workbook = Workbook() 
sheet0 = workbook.create_sheet(index=0) # 创建sheet0
sheet0.column_dimensions['A'].width=15 # 设置A列宽度
sheet0.column_dimensions['B'].width=22 # 设置B列宽度
sheet0.column_dimensions['C'].width=15 # 设置C列宽度
sheet0.column_dimensions['D'].width=15 # 设置D列宽度
# 循环写入数据，居中对齐
for i in range(len(a)):
    sheet0.cell(i+1,1).value = a[i] # 写入数据

for j in range(len(b)):
    sheet0.cell(j+1,2).value = b[j] # 写入数据

workbook.save('test.xlsx') # 保存文件